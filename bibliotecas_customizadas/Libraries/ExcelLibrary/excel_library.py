import openpyxl
from robot.api.deco import keyword

"""
Biblioteca personalizada para manipulação de planilhas Excel usando openpyxl.

Esta biblioteca fornece keywords para interagir com planilhas Excel no formato .xlsx.
Ela permite carregar planilhas, buscar dados por tamanho, raça e temperamento, além de contar animais por temperamento.

Dependência:
    openpyxl: Biblioteca do Python para ler e escrever arquivos Excel (.xlsx).

Funções:
    - Carregar Planilha: Carrega a planilha e retorna o objeto WorkBook.
    - Buscar Animais Por Tamanho: Retorna uma lista de animais que possuem o tamanho especificado.
    - Buscar Animal Por Raça: Retorna os detalhes de uma raça específica.
    - Contar Animais Por Temperamento: Conta quantos animais possuem o temperamento especificado.

Exemplo de uso:
    Exemplo de uso:
    arquivo = "./bibliotecas_customizadas/tests/my_custom_libraries_excel/exem_planilha_testes_automatizados_2.xlsx"  # Substitua pelo caminho real do seu arquivo
    planilha_dados = carregar_planilha(arquivo)

    Chamando as funções
    print("\nBuscar Animais Por Tamanho:")
    print(buscar_por_tamanho(planilha_dados, "pequeno"))
    print("\nBuscar Animais Por Raça:")
    print(buscar_por_raca(planilha_dados, "Pinscher"))
    print("\nContar Animais Por Temperamento:")
    print(contar_por_temperamento(planilha_dados, "Atento"), "\n")
"""

@keyword("Carregar Planilha")
def carregar_planilha(arquivo):
    """Carrega a planilha e retorna o objeto WorkBook."""
    return openpyxl.load_workbook(arquivo)

@keyword("Buscar Animais Por Tamanho")
def buscar_por_tamanho(WorkBook, tamanho):
    """Retorna uma lista de animais que possuem o tamanho especificado."""
    ws = WorkBook.active  # Obtém a aba ativa
    resultados = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[2].lower() == tamanho.lower():
            resultados.append(row)
    return resultados

@keyword("Buscar Animal Por Raça")
def buscar_por_raca(WorkBook, raca):
    """Retorna os detalhes de uma raça específica."""
    ws = WorkBook.active  # Obtém a aba ativa
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[1].lower() == raca.lower():
            return row
    return "Raça não encontrada"

@keyword("Contar Animais Por Temperamento")
def contar_por_temperamento(WorkBook, temperamento):
    """Conta quantos animais possuem o temperamento especificado."""
    ws = WorkBook.active  # Obtém a aba ativa
    contador = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[3].lower() == temperamento.lower():
            contador += 1
    return contador